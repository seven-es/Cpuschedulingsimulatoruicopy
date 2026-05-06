#!/usr/bin/env python3
"""
run_all.py
==========
Watches algorithms/workload.txt for changes, then:
  1. Compiles all 4 C algorithms (gcc required).
  2. Runs each executable — each writes 3 CSV files:
       {ALGO}_gantt_chart.csv
       {ALGO}_process_stats.csv
       {ALGO}_summary.csv
  3. Converts every CSV to a formatted Excel (.xlsx) file (needs openpyxl).
  4. Merges all results into public/all_results.json for the React UI.

Install openpyxl once:  pip install openpyxl
Run:  python run_all.py
"""

import os
import sys
import subprocess
import json
import time
import csv
import shutil
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("WARNING: openpyxl not installed — Excel (.xlsx) files won't be created.")
    print("         Install with:  pip install openpyxl\n")

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).parent.resolve()
ALGO_DIR      = SCRIPT_DIR / "algorithms"
PUBLIC_DIR    = SCRIPT_DIR / "public"
WORKLOAD_FILE = ALGO_DIR / "workload.txt"

IS_WIN = sys.platform == "win32"
EXE_SUFFIX = ".exe" if IS_WIN else ""

ALGORITHMS = {
    "FCFS":     "FCFS.c",
    "RR":       "RR.c",
    "Priority": "PriorityQueue.c",
    "SRFJ":     "SRFJ.c",
}

ALGO_LABELS = {
    "FCFS":     "First Come First Serve",
    "RR":       "Round Robin",
    "Priority": "Priority (Preemptive)",
    "SRFJ":     "Shortest Remaining First Job",
}

# ── Compilation ───────────────────────────────────────────────────────────────
def find_gcc():
    for name in ("gcc", "gcc.exe", "cc"):
        if shutil.which(name):
            return name
    return None

def compile_all():
    gcc = find_gcc()
    if not gcc:
        print("ERROR: gcc not found. Install MinGW-w64 (Windows) or build-essential (Linux/Mac).")
        sys.exit(1)

    compiled = {}
    for name, src in ALGORITHMS.items():
        src_path = ALGO_DIR / src
        exe_path = ALGO_DIR / (name + EXE_SUFFIX)
        result = subprocess.run(
            [gcc, str(src_path), "-o", str(exe_path), "-lm"],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            compiled[name] = exe_path
            print(f"  [OK]   Compiled {name}")
        else:
            print(f"  [FAIL] {name}: {result.stderr.strip()}")
    return compiled

# ── Running ───────────────────────────────────────────────────────────────────
def run_algorithm(exe_path, algo_name):
    result = subprocess.run(
        [str(exe_path)],
        capture_output=True, text=True,
        cwd=str(ALGO_DIR)
    )
    if result.returncode == 0:
        print(f"  [OK]   Ran {algo_name}")
        return True
    else:
        print(f"  [FAIL] {algo_name}: {result.stderr.strip() or result.stdout.strip()}")
        return False

# ── CSV helpers ───────────────────────────────────────────────────────────────
def read_csv(filepath):
    if not filepath.exists():
        return []
    rows = []
    with open(filepath, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    return rows

# ── Excel output ──────────────────────────────────────────────────────────────
HEADER_BG   = "1E40AF"  # deep blue
HEADER_FONT = "FFFFFF"
ALT_BG      = "EFF6FF"  # light blue

def csv_to_excel(csv_path, xlsx_path, sheet_title):
    if not HAS_EXCEL or not csv_path.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    hfill = PatternFill("solid", fgColor=HEADER_BG)
    hfont = Font(color=HEADER_FONT, bold=True, name="Calibri", size=11)
    afill = PatternFill("solid", fgColor=ALT_BG)
    thin  = Side(style="thin", color="BFDBFE")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for ri, row in enumerate(reader):
            ws.append(row)
            for ci, cell in enumerate(ws[ri + 1], start=1):
                cell.border = bdr
                cell.alignment = center
                if ri == 0:
                    cell.fill = hfill
                    cell.font = hfont
                elif ri % 2 == 1:
                    cell.fill = afill

    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = max(width + 4, 12)

    wb.save(xlsx_path)

# ── JSON for UI ───────────────────────────────────────────────────────────────
def safe_int(v, default=0):
    try: return int(v)
    except: return default

def safe_float(v, default=0.0):
    try: return float(v)
    except: return default

def build_ui_json(algo_name):
    gantt_csv   = ALGO_DIR / f"{algo_name}_gantt_chart.csv"
    stats_csv   = ALGO_DIR / f"{algo_name}_process_stats.csv"
    summary_csv = ALGO_DIR / f"{algo_name}_summary.csv"

    gantt_rows   = read_csv(gantt_csv)
    stats_rows   = read_csv(stats_csv)
    summary_rows = read_csv(summary_csv)

    gantt = [
        {"pid": r["PID"], "start": safe_int(r["Start"]), "end": safe_int(r["End"])}
        for r in gantt_rows
    ]

    stats = []
    for r in stats_rows:
        stats.append({
            "pid":           r.get("PID", ""),
            "arrival":       safe_int(r.get("Arrival", 0)),
            "priority":      safe_int(r.get("Priority", 0)),
            # New format uses CPUBurst/IOBurst; old format used Burst
            "cpuBurst":      safe_int(r.get("CPUBurst", r.get("Burst", 0))),
            "ioBurst":       safe_int(r.get("IOBurst", 0)),
            "waitingTime":   safe_int(r.get("Waiting", 0)),
            "turnaroundTime":safe_int(r.get("Turnaround", 0)),
            "responseTime":  safe_int(r.get("Response", 0)),
            "finish":        safe_int(r.get("Finish", 0)),
        })

    summary = {r["Metric"]: r["Value"] for r in summary_rows}

    return {
        "algorithm": algo_name,
        "label":     ALGO_LABELS.get(algo_name, algo_name),
        "timestamp": int(time.time()),
        "ganttChart":   gantt,
        "processStats": stats,
        "summary":      summary,
    }

# ── Main pipeline ─────────────────────────────────────────────────────────────
def run_all(compiled):
    print(f"\n[{time.strftime('%H:%M:%S')}] Processing workload.txt ...")

    all_data = {}
    for algo_name, exe_path in compiled.items():
        if not run_algorithm(exe_path, algo_name):
            continue

        # CSV → Excel
        for suffix, title in [
            ("gantt_chart",   "Gantt Chart"),
            ("process_stats", "Process Stats"),
            ("summary",       "Summary"),
        ]:
            csv_p  = ALGO_DIR / f"{algo_name}_{suffix}.csv"
            xlsx_p = ALGO_DIR / f"{algo_name}_{suffix}.xlsx"
            csv_to_excel(csv_p, xlsx_p, title)
            if HAS_EXCEL and xlsx_p.exists():
                print(f"  [XL]   {xlsx_p.name}")

        # Build UI JSON
        data = build_ui_json(algo_name)
        all_data[algo_name] = data

    # Write combined JSON for React UI
    PUBLIC_DIR.mkdir(exist_ok=True)
    combined = {"timestamp": int(time.time()), "algorithms": all_data}
    out = PUBLIC_DIR / "all_results.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(combined, f, indent=2)
    print(f"  [UI]   Written → {out}")

def main():
    print("=" * 60)
    print("  CPU Scheduling Simulator — run_all.py")
    print("=" * 60)
    print(f"  Algorithms dir : {ALGO_DIR}")
    print(f"  Workload file  : {WORKLOAD_FILE}")
    print(f"  UI output dir  : {PUBLIC_DIR}")
    print()

    if not ALGO_DIR.exists():
        print(f"ERROR: algorithms/ directory not found at {ALGO_DIR}")
        sys.exit(1)

    if not WORKLOAD_FILE.exists():
        print("No workload.txt found — creating default one ...")
        with open(WORKLOAD_FILE, "w") as f:
            f.write("# CPU Scheduling Workload\n")
            f.write("# Format: PID  Arrival  Burst  Priority\n")
            f.write("# Lower priority number = higher urgency\n")
            f.write("# quantum=2\n")
            f.write("P1 0 8 2\n")
            f.write("P2 1 4 1\n")
            f.write("P3 2 9 3\n")
            f.write("P4 3 5 2\n")

    print("Compiling algorithms ...")
    compiled = compile_all()
    if not compiled:
        print("ERROR: No algorithms compiled. Exiting.")
        sys.exit(1)

    # Run once immediately
    run_all(compiled)

    # Watch for changes
    print(f"\nWatching {WORKLOAD_FILE.name} for changes (Ctrl+C to stop) ...\n")
    last_mtime = WORKLOAD_FILE.stat().st_mtime if WORKLOAD_FILE.exists() else 0

    while True:
        try:
            time.sleep(1)
            mtime = WORKLOAD_FILE.stat().st_mtime if WORKLOAD_FILE.exists() else 0
            if mtime != last_mtime:
                last_mtime = mtime
                run_all(compiled)
        except KeyboardInterrupt:
            print("\nStopped.")
            break
        except Exception as exc:
            print(f"  [ERR] {exc}")

if __name__ == "__main__":
    main()
